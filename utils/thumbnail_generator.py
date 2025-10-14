import os
import boto3
from PIL import Image, ImageDraw, ImageFont
from io import BytesIO
import fitz  # PyMuPDF for PDF thumbnails
from openpyxl import load_workbook
from docx import Document
import pandas as pd
from werkzeug.utils import secure_filename
import mimetypes

# Thumbnail settings
THUMBNAIL_SIZE = (200, 200)
THUMBNAIL_QUALITY = 85


def get_file_type(filename):
    """Determine file type from filename"""
    extension = filename.rsplit(".", 1)[1].lower() if "." in filename else ""

    image_types = {"jpg", "jpeg", "png", "gif", "bmp", "tiff"}
    document_types = {"pdf", "docx", "doc"}
    spreadsheet_types = {"xlsx", "xls", "csv"}
    text_types = {"txt"}

    if extension in image_types:
        return "image"
    elif extension in document_types:
        return "document"
    elif extension in spreadsheet_types:
        return "spreadsheet"
    elif extension in text_types:
        return "text"
    else:
        return "unknown"


def create_text_thumbnail(
    text, title="Document", bg_color=(255, 255, 255), text_color=(50, 50, 50)
):
    """Create a thumbnail with text content preview"""
    img = Image.new("RGB", THUMBNAIL_SIZE, bg_color)
    draw = ImageDraw.Draw(img)

    try:
        # Try to use a nice font
        title_font = ImageFont.truetype("arial.ttf", 14)
        text_font = ImageFont.truetype("arial.ttf", 10)
    except:
        # Fallback to default font
        title_font = ImageFont.load_default()
        text_font = ImageFont.load_default()

    # Draw title
    draw.text((10, 10), title, fill=text_color, font=title_font)

    # Draw text preview (first few lines)
    y_offset = 35
    lines = text.split("\n")[:8]  # First 8 lines
    for line in lines:
        if y_offset > THUMBNAIL_SIZE[1] - 20:
            break
        # Truncate long lines
        if len(line) > 25:
            line = line[:22] + "..."
        draw.text((10, y_offset), line, fill=text_color, font=text_font)
        y_offset += 15

    return img


def create_icon_thumbnail(icon_text, bg_color, text_color=(255, 255, 255)):
    """Create a simple icon-style thumbnail"""
    img = Image.new("RGB", THUMBNAIL_SIZE, bg_color)
    draw = ImageDraw.Draw(img)

    try:
        font = ImageFont.truetype("arial.ttf", 24)
    except:
        font = ImageFont.load_default()

    # Center the icon text
    bbox = draw.textbbox((0, 0), icon_text, font=font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]

    x = (THUMBNAIL_SIZE[0] - text_width) // 2
    y = (THUMBNAIL_SIZE[1] - text_height) // 2

    draw.text((x, y), icon_text, fill=text_color, font=font)
    return img


def generate_pdf_thumbnail(file_content):
    """Generate thumbnail from PDF first page"""
    try:
        pdf_document = fitz.open(stream=file_content, filetype="pdf")
        if len(pdf_document) > 0:
            page = pdf_document[0]
            # Render page to image
            mat = fitz.Matrix(1.0, 1.0)  # zoom factor
            pix = page.get_pixmap(matrix=mat)
            img_data = pix.tobytes("png")

            img = Image.open(BytesIO(img_data))
            img.thumbnail(THUMBNAIL_SIZE, Image.Resampling.LANCZOS)

            # Create a new image with white background
            thumbnail = Image.new("RGB", THUMBNAIL_SIZE, (255, 255, 255))
            # Center the image
            x = (THUMBNAIL_SIZE[0] - img.width) // 2
            y = (THUMBNAIL_SIZE[1] - img.height) // 2
            thumbnail.paste(img, (x, y))

            pdf_document.close()
            return thumbnail
    except Exception as e:
        print(f"Error generating PDF thumbnail: {e}")

    return create_icon_thumbnail("PDF", (220, 53, 69))


def generate_docx_thumbnail(file_content):
    """Generate thumbnail from DOCX content"""
    try:
        doc = Document(BytesIO(file_content))
        text_content = ""

        for paragraph in doc.paragraphs[:10]:  # First 10 paragraphs
            if paragraph.text.strip():
                text_content += paragraph.text + "\n"
                if len(text_content) > 300:  # Limit text length
                    break

        if text_content.strip():
            return create_text_thumbnail(
                text_content, "Word Document", (41, 128, 185), (255, 255, 255)
            )
    except Exception as e:
        print(f"Error generating DOCX thumbnail: {e}")

    return create_icon_thumbnail("DOCX", (41, 128, 185))


def generate_excel_thumbnail(file_content, filename):
    """Generate thumbnail from Excel/CSV content"""
    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(BytesIO(file_content), nrows=10)
            title = "CSV File"
            bg_color = (40, 167, 69)
        else:
            wb = load_workbook(BytesIO(file_content))
            ws = wb.active

            # Get first few rows and columns
            data = []
            for row in ws.iter_rows(max_row=6, max_col=3, values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                data.append(" | ".join(row_data))

            text_content = "\n".join(data)
            title = "Excel File"
            bg_color = (40, 167, 69)

        if "text_content" in locals():
            return create_text_thumbnail(text_content, title, bg_color, (255, 255, 255))
        else:
            # For CSV, create a simple preview
            preview_text = df.head(5).to_string(index=False, max_cols=2)
            return create_text_thumbnail(preview_text, title, bg_color, (255, 255, 255))

    except Exception as e:
        print(f"Error generating Excel/CSV thumbnail: {e}")

    return create_icon_thumbnail("XLS", (40, 167, 69))


def generate_text_thumbnail(file_content):
    """Generate thumbnail from text file"""
    try:
        text = file_content.decode("utf-8")[:500]  # First 500 characters
        return create_text_thumbnail(
            text, "Text File", (108, 117, 125), (255, 255, 255)
        )
    except Exception as e:
        print(f"Error generating text thumbnail: {e}")

    return create_icon_thumbnail("TXT", (108, 117, 125))


def generate_image_thumbnail(file_content):
    """Generate thumbnail from image file"""
    try:
        img = Image.open(BytesIO(file_content))
        # Convert to RGB if necessary
        if img.mode in ("RGBA", "LA", "P"):
            background = Image.new("RGB", img.size, (255, 255, 255))
            if img.mode == "P":
                img = img.convert("RGBA")
            background.paste(img, mask=img.split()[-1] if img.mode == "RGBA" else None)
            img = background
        elif img.mode != "RGB":
            img = img.convert("RGB")

        img.thumbnail(THUMBNAIL_SIZE, Image.Resampling.LANCZOS)

        # Create a new image with white background and center the thumbnail
        thumbnail = Image.new("RGB", THUMBNAIL_SIZE, (255, 255, 255))
        x = (THUMBNAIL_SIZE[0] - img.width) // 2
        y = (THUMBNAIL_SIZE[1] - img.height) // 2
        thumbnail.paste(img, (x, y))

        return thumbnail
    except Exception as e:
        print(f"Error generating image thumbnail: {e}")

    return create_icon_thumbnail("IMG", (255, 193, 7))


def generate_thumbnail(file_content, filename):
    """Main function to generate thumbnail based on file type"""
    file_type = get_file_type(filename)

    if file_type == "image":
        return generate_image_thumbnail(file_content)
    elif file_type == "document":
        if filename.lower().endswith(".pdf"):
            return generate_pdf_thumbnail(file_content)
        elif filename.lower().endswith((".docx", ".doc")):
            return generate_docx_thumbnail(file_content)
    elif file_type == "spreadsheet":
        return generate_excel_thumbnail(file_content, filename)
    elif file_type == "text":
        return generate_text_thumbnail(file_content)

    # Default fallback
    return create_icon_thumbnail("FILE", (128, 128, 128))


def save_thumbnail_to_s3(thumbnail_img, s3_client, bucket, s3_key):
    """Save thumbnail image to S3"""
    # Ensure s3_key and bucket are strings (not bytes)
    if isinstance(s3_key, bytes):
        s3_key = s3_key.decode('utf-8')
    if isinstance(bucket, bytes):
        bucket = bucket.decode('utf-8')

    thumbnail_buffer = BytesIO()
    thumbnail_img.save(thumbnail_buffer, format="JPEG", quality=THUMBNAIL_QUALITY)
    thumbnail_buffer.seek(0)

    # Generate thumbnail key by replacing the last extension
    if "." in s3_key:
        base_name = s3_key.rsplit(".", 1)[0]
        thumbnail_key = f"{base_name}_thumb.jpg"
    else:
        thumbnail_key = f"{s3_key}_thumb.jpg"

    s3_client.upload_fileobj(
        thumbnail_buffer,
        bucket,
        thumbnail_key,
        ExtraArgs={
            "ContentType": "image/jpeg",
            "CacheControl": "max-age=31536000",  # 1 year cache
        },
    )

    return thumbnail_key


def save_thumbnail_locally(thumbnail_img, local_path):
    """Save thumbnail image locally"""
    thumbnail_path = local_path.rsplit(".", 1)[0] + "_thumb.jpg"
    os.makedirs(os.path.dirname(thumbnail_path), exist_ok=True)
    thumbnail_img.save(thumbnail_path, format="JPEG", quality=THUMBNAIL_QUALITY)
    return thumbnail_path
